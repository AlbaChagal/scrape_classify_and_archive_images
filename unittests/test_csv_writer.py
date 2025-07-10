# External Imports
from openpyxl import load_workbook
import os
from PIL import Image as PILImage
import tempfile
import shutil
import unittest

# Internal Imports
from archive.csv_writer import CSVWriter


class TestCSVWriter(unittest.TestCase):

    def test_init_valid(self):
        writer = CSVWriter("test.xlsx")
        self.assertEqual(writer.csv_file, "test.xlsx")
        self.assertEqual(writer.ws.title, "Sheet")

    def test_create_sheet_from_images_real_images(self):
        with tempfile.TemporaryDirectory() as tmpdirname:
            # Create real image files
            image_filenames = ["img1.jpg", "img2.png"]
            for fname in image_filenames:
                img = PILImage.new('RGB', (100, 100), color='red')
                img.save(os.path.join(tmpdirname, fname))

            # Add one non-image file
            with open(os.path.join(tmpdirname, "not_image.txt"), 'w') as f:
                f.write("Just some text")

            # Output Excel file path
            output_file = os.path.join(tmpdirname, "test_output.xlsx")

            writer = CSVWriter(output_file)
            writer.create_sheet_from_images(tmpdirname, title="Images")

            # Check the file exists
            self.assertTrue(os.path.exists(output_file))

            # Check workbook contents
            wb = load_workbook(output_file)
            ws = wb["Images"]
            self.assertEqual(ws.cell(row=1, column=1).value, "img2")
            self.assertEqual(ws.cell(row=2, column=1).value, "img1")
            shutil.rmtree(tmpdirname)

    def test_create_sheet_from_images_invalid_path(self):
        writer = CSVWriter("dummy.xlsx")
        with self.assertRaises(AssertionError):
            writer.create_sheet_from_images('123')  # Non-string path

        with self.assertRaises(AssertionError):
            writer.create_sheet_from_images("nonexistent_path")  # Not a directory

        with self.assertRaises(AssertionError):
            os.makedirs("temp_dir", exist_ok=True)
            writer.create_sheet_from_images("temp_dir", title=123)  # Title is not a string
        os.rmdir("temp_dir")
        os.remove(writer.csv_file)


if __name__ == "__main__":
    unittest.main()

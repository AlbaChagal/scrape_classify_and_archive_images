# External Imports
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
import os
from typing import Tuple, Optional


class CSVWriter:
    def __init__(self, csv_file: str, ):
        """
        A class writing CSV files
        :param csv_file: The path to the output CSV file
        """
        assert type(csv_file) is str, f'csv_file is not a string, got {type(csv_file)}'
        self.csv_file: str = csv_file
        self.wb: Workbook = Workbook()
        self.ws: Worksheet = self.wb.active

    @staticmethod
    def get_allowed_image_formats() -> Tuple[str, str, str, str, str]:
        """
        Returns the allowed formats for the CSVWriter
        :return: Allowed formats
        """
        return 'jpg', 'jpeg', 'png', 'bmp', 'gif'

    def create_sheet_from_images(self, images_path: str, title: Optional[str] = None):
        """
        Creates a sheet from a folder of images.
        Where column A has the name of the file and column B has the image and each row is an image
        :param images_path: The directory where the images are located.
        :param title: The title of the sheet (Optional)
        :return: None. saves the sheet to self.csv_file
        """

        assert type(images_path) is str, \
            f'images_path is not a string, got {type(images_path)}'
        assert os.path.isdir(images_path), \
            f'images_path is not a directory, got {images_path}'
        assert type(title) is str, \
            f'title is not a string, got {type(title)}'
        
        # Set the title if needed
        if title:
            self.ws.title = title

        # Declare all loop-variable types once in advance (for Cythonization)
        image_path: str
        img_openpyxl: Image

        # Iterate over the images in the folder and add them to the workbook
        row: int = 1  # Start from the first row
        for filename in os.listdir(images_path):
            if filename.lower().endswith(self.get_allowed_image_formats()):
                # Set image name in Column A
                self.ws.cell(row=row, column=1, value=filename.split('.')[0])

                # Get the full path to the image file
                image_path = os.path.join(images_path, filename)

                # Create an Image object for openpyxl
                img_openpyxl = Image(image_path)
                img_openpyxl.width = 150  # Resize image to fit in the cell (optional)
                img_openpyxl.height = 150
        
                # Insert the image into Column B
                self.ws.add_image(img_openpyxl, f'B{row}')

                # Move to the next row
                row += 1

        # Save the workbook to a file
        self.wb.save(self.csv_file)

        print(f"Excel file created: {self.csv_file}")

if __name__ == "__main__":
    # Folder where your images are located
    folder_path_main = "../data/archive_images"  # Replace with your actual folder path

    writer = CSVWriter("outputs/idx_images.xlsx")
    writer.create_sheet_from_images(folder_path_main)